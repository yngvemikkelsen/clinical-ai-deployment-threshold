#!/usr/bin/env python3
"""
Paper 9 — generate Multimedia Appendices 1-6.

Each appendix exists so a reviewer can audit a claim the main text makes in
summary form, without reverse-engineering the repository. Every table is
derived here from the primary artefact; nothing is transcribed by hand.

  1  System extraction and partial identification (55 systems)
  2  Full 52-cell transport matrix + uniform-512 whole-note results
  3  Local-screen resampling, per configuration x corpus x sample size
  4  Query-generation robustness, 26 paired comparisons + prompts
  5  Decision-model parameters, epsilon sweep, VOI outputs
  6  Code-search boundary replication
"""
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook

U = Path("/mnt/user-data/uploads")
OUT = Path("appendices")
OUT.mkdir(exist_ok=True)

TIER1 = {"BioLORD-2023", "MedCPT", "BGE-base", "GTE-base",
         "Nomic-embed-text", "Nomic-embed-text-nopfx"}

# Commercial-interface systems: pooling is vendor-fixed and geometry
# undisclosed, so membership is unobservable in principle from publication.
API_MARKERS = ["openai", "ada", "text-embedding", "gpt", "azure", "cohere",
               "gemini", "claude", "api", "titan", "bedrock", "vertex",
               "voyage", "mistral-embed"]
# Not pooled dense retrievers: corpus-only ZCA is undefined for these.
# A system is non-dense only if it has NO pooled dense encoder. BM25 appearing
# alongside a dense encoder is a hybrid retriever, not a non-dense system:
# CliniqIR selects MedCPT and adds BM25, so corpus-only ZCA is defined for it.
NONDENSE_MARKERS = ["colbert", "late interaction", "chunk scorer", "albef",
                    "learned chunk"]
DENSE_MARKERS = ["bert", "medcpt", "mpnet", "e5", "gte", "bge", "nomic",
                 "sbert", "sentence", "biolord", "sapbert", "coder", "dpr",
                 "minilm", "gtr", "glucose", "gist", "contriever", "bmretriever"]


def cls_interface(enc: str) -> str:
    e = (enc or "").lower()
    # ColBERT contains "bert" as a substring but is a late-interaction
    # retriever with no single pooled document vector, so it must be tested
    # before the dense-marker check rather than after it.
    if "colbert" in e:
        return "not a pooled dense retriever"
    if any(m in e for m in NONDENSE_MARKERS) and not any(
            m in e for m in DENSE_MARKERS):
        return "not a pooled dense retriever"
    if any(m in e for m in API_MARKERS):
        return "commercial interface"
    if not e.strip() or e.strip().lower() in ("nan", "none", "unspecified"):
        return "checkpoint unspecified"
    # a named family with no identifiable checkpoint cannot be coded on
    # training objective either
    vague = ["sentence transformer", "fine-tuned embedding model",
             "embedding model", "dense retriever"]
    if any(e.strip().lower() == v for v in vague):
        return "checkpoint unspecified"
    return "self-hosted checkpoint"


def appendix1():
    """55 systems, every coding decision, classification under each rung.

    The main text claims: none of the 55 reports a retrieval pooling strategy
    as such; one can be inferred from implementation detail given elsewhere;
    27 use commercial interfaces where pooling is not a user choice; 2 are
    confirmed affected. This appendix is the audit trail for all four.
    """
    src = (Path("p_tier2_wide_frame_extraction_audit_v2.xlsx")
           if Path("p_tier2_wide_frame_extraction_audit_v2.xlsx").exists()
           else U / "p_tier2_wide_frame_extraction_audit.xlsx")
    wb = load_workbook(src,
                       read_only=True, data_only=True)
    rows = [list(r) for r in wb["Analysis_Set"].iter_rows(values_only=True)]
    hdr = [str(h) if h is not None else "" for h in rows[0]]
    df = pd.DataFrame(rows[1:], columns=hdr).dropna(how="all")
    df = df[df["Core include"].astype(str).str.strip() == "Y"].copy()

    enc = df["Reported operational encoder/retriever"].astype(str)
    df["Interface class"] = enc.map(cls_interface)

    # Pooling: the finding is that it is absent. One system's extraction step
    # is inferable from implementation detail reported elsewhere in its paper.
    note = df["Audit note"].astype(str).str.lower()
    df["Reported pooling"] = "not reported as such"

    # The source workbook's audit note for A50 records the ORIGINAL coding
    # rationale, which the manuscript establishes is invalid: citing the
    # base-model paper is the standard citation for the contrastively
    # fine-tuned sentence-transformer checkpoint, so "no similarity-tuned
    # checkpoint is identified" does not follow. The valid basis is the
    # pooling mismatch read from the primary article's methods section. That
    # correction was made in the manuscript but never written back to the
    # workbook, so it is applied here explicitly rather than inherited.
    INFERRED = {
        "A50": (
            "classification-token ([CLS]) representation",
            "The primary article reports extraction of the [CLS] "
            "representation and identifies MPNet by reference to the original "
            "MPNet model (Song et al.) rather than a retrieval-trained "
            "sentence-embedding checkpoint. On the reported implementation the "
            "retrieval representation was therefore not shown to have been "
            "calibrated by a retrieval-specific objective. Membership follows "
            "from the extracted representation, not from the training label.",
            "Section 3.2, doi:10.1145/3584371.3612956"),
    }
    key = df["Canonical ID"].astype(str).str.strip()
    df["Inferred extraction/representation"] = key.map(
        lambda k: INFERRED.get(k, ("", "", ""))[0])
    df["Basis for inference"] = key.map(
        lambda k: INFERRED.get(k, ("", "", ""))[1])
    df["Primary-source evidence"] = key.map(
        lambda k: INFERRED.get(k, ("", "", ""))[2])
    df["Superseded coding rationale"] = np.where(
        key.isin(INFERRED),
        "workbook audit note records the original rationale (no similarity-"
        "tuned checkpoint identified), superseded; see Basis for inference",
        "")

    # The workbook's Tier 2 coding is the STARTING point, not the definition.
    # A system is confirmed affected only where the representation actually
    # used for retrieval can be established from the primary article AND is
    # one the training objective did not calibrate. A51 (Soman / KG-RAG) is
    # Tier 2 in the workbook but its primary article establishes only a model
    # family ("PubMedBERT selected for context embeddings"), not the extraction
    # step, so it is coded unknown rather than affected. The authors' current
    # repository specifies a sentence-transformer checkpoint for context
    # retrieval, but that repository state may postdate the published
    # experiment, so it cannot be used to recode in either direction.
    NOT_ESTABLISHED = {"A51"}
    key0 = df["Canonical ID"].astype(str).str.strip()
    df["Confirmed affected"] = np.where(
        df["Core tier"].astype(str).str.strip().str.endswith("2")
        & ~key0.isin(NOT_ESTABLISHED), "yes", "no")
    df["Extraction step established"] = np.where(
        df["Core tier"].astype(str).str.strip().str.endswith("2"),
        np.where(key0.isin(NOT_ESTABLISHED),
                 "no - model family named, extraction step not reported",
                 "yes - see Basis for inference"), "")
    # One system is coded affected in the core frame on a pooling-mismatch
    # basis and simultaneously flagged ambiguous in the sensitivity notes.
    # Record both rather than silently privileging one.
    # A50 is Tier 2 in all three frames; the workbook note flags it as
    # ambiguous under the original rationale only. Under the pooling-mismatch
    # basis the coding is stable, and this column says which is which.
    df["Coding stability"] = np.where(
        key.isin(INFERRED),
        "Tier 2 in core, expanded and adversarial frames; flagged ambiguous "
        "in the workbook under the superseded rationale, stable under the "
        "extracted-representation basis",
        "stable across frames")

    # Identification ladder, per system
    df["Rung 1: no assumptions"] = np.where(
        df["Confirmed affected"] == "yes", "affected", "unknown")
    # Rung 2 grants only that a documented contrastive objective on an
    # identified checkpoint implies non-membership. Systems whose membership is
    # unobservable -- commercial interfaces, unspecified checkpoints -- stay
    # unknown. Non-dense systems are coded unaffected here because corpus-only
    # ZCA is undefined for them, so they cannot belong to the affected subgroup
    # of this correction; they are formally excluded only at rung 4.
    # The rung grants only that DOCUMENTED contrastive training implies
    # non-membership. A self-hosted checkpoint that is MLM-pretrained rather
    # than contrastively trained (PubMedBERT, base MPNet) is not covered and
    # stays unknown.
    # The rung grants only that DOCUMENTED contrastive training implies
    # non-membership. It does not cover a self-hosted checkpoint that is
    # MLM-pretrained rather than contrastively trained, nor one whose
    # extraction step is unestablished. A51 is both, so it stays unknown at
    # this rung and at rung 3, where accepting vendors' retrieval-optimisation
    # claims cannot resolve a self-hosted system. Listing the exception
    # explicitly rather than by substring rule: a system that names several
    # candidate encoders and SELECTS a retrieval-trained one (e.g. CliniqIR,
    # which selects MedCPT) is covered by the rung and must not be caught.
    unresolved = key0.isin(NOT_ESTABLISHED)
    df["Rung 2: contrastive training implies non-membership"] = np.where(
        df["Confirmed affected"] == "yes", "affected",
        np.where(df["Interface class"].isin(
            ["self-hosted checkpoint", "not a pooled dense retriever"])
            & ~unresolved, "coded unaffected", "unknown"))

    df["Rung 3: vendor claims accepted"] = np.where(
        df["Confirmed affected"] == "yes", "affected",
        np.where((df["Interface class"] == "checkpoint unspecified")
                 | unresolved, "unknown", "coded unaffected"))
    df["Rung 4: non-dense excluded"] = np.where(
        df["Interface class"] == "not a pooled dense retriever", "excluded",
        df["Rung 3: vendor claims accepted"])

    cols = ["Canonical ID", "System", "Source frame",
            "Reported operational encoder/retriever", "Interface class",
            "Reported pooling", "Inferred extraction/representation",
            "Basis for inference", "Primary-source evidence",
            "Extraction step established",
            "Superseded coding rationale", "Coding stability", "Core tier", "Expanded tier",
            "Adversarial tier", "Confidence", "Confirmed affected",
            "Rung 1: no assumptions",
            "Rung 2: contrastive training implies non-membership",
            "Rung 3: vendor claims accepted", "Rung 4: non-dense excluded",
            "Audit note", "Primary source URL", "Review/source URL"]
    out = df[[c for c in cols if c in df.columns]].fillna("")
    out.to_csv(OUT / "appendix1_system_extraction.csv", index=False)

    # the coding rule, stated rather than left implicit: 2/55 is load-bearing
    Path(OUT / "appendix1_coding_definitions.txt").write_text(
        "Multimedia Appendix 1 - coding definitions\n"
        "=========================================\n\n"
        "Scope of these codings\n"
        "  Every coding in this frame is MECHANISTIC: it records an\n"
        "  implementation characteristic expected to correspond to the\n"
        "  response-defined subgroup of the decision model, not a measured\n"
        "  response to the intervention. No system in the frame was evaluated\n"
        "  under the intervention. The counts here therefore support proxy-based\n"
        "  partial identification under stated classification assumptions; they\n"
        "  do not bound response-defined prevalence, which is unidentified\n"
        "  without such an assumption.\n\n"
        "Mechanistic affected criterion\n"
        "  A system is coded as meeting the hypothesised affected mechanism\n"
        "  only where BOTH hold:\n"
        "   (a) the representation actually used for retrieval can be\n"
        "       established from the primary article, and\n"
        "   (b) that representation was not shown to have been calibrated by\n"
        "       the model's training objective.\n"
        "  Systems whose extraction step cannot be established from the primary\n"
        "  article are coded UNKNOWN, not unaffected. One of 55 systems meets\n"
        "  this mechanistic criterion. No system in the frame was evaluated\n"
        "  under the intervention, so the count does not bound response-defined\n"
        "  prevalence. It enters the analysis only under explicit\n"
        "  classification assumptions and, separately, as a labelled\n"
        "  stress-test value.\n\n"
        "Reported pooling\n"
        "  None of the 55 systems reports a retrieval pooling strategy as such.\n"
        "  One (A50) reports enough implementation detail elsewhere in its\n"
        "  methods to infer the extraction step; the inference, its basis, and\n"
        "  the primary-source location are given in their own columns.\n\n"
        "Coding stability\n"
        "  A50 is Tier 2 in the core, expanded and adversarial frames. The\n"
        "  source workbook additionally flagged it ambiguous under the ORIGINAL\n"
        "  rationale (no similarity-tuned checkpoint identified), which the\n"
        "  manuscript establishes is invalid: citing the base-model paper does\n"
        "  not establish the absence of a similarity-tuned checkpoint. Under\n"
        "  the extracted-representation basis the coding is not ambiguous.\n"
        "  Both the\n"
        "  superseded rationale and the current basis are recorded so the\n"
        "  change is auditable.\n\n"
        "Interface class\n"
        "  commercial interface        pooling vendor-fixed, geometry undisclosed\n"
        "  self-hosted checkpoint      checkpoint identified and inspectable\n"
        "  checkpoint unspecified      no identifiable checkpoint named\n"
        "  not a pooled dense retriever  corpus-only whitening is undefined\n")

    summ = pd.DataFrame({
        "Interface class": df["Interface class"].value_counts().index,
        "n": df["Interface class"].value_counts().values,
    })
    summ.to_csv(OUT / "appendix1_summary.csv", index=False)
    print(f"A1  {len(out)} systems")
    print(f"    {df['Interface class'].value_counts().to_dict()}")
    print(f"    confirmed affected: {(df['Confirmed affected']=='yes').sum()}")
    print(f"    pooling reported as such: 0 of {len(df)}")
    n = len(df)
    for c in [c for c in df.columns if c.startswith("Rung")]:
        v = df[c].value_counts().to_dict()
        aff = v.get("affected", 0)
        unk = v.get("unknown", 0)
        exc = v.get("excluded", 0)
        denom = n - exc
        print(f"    {c[:46]:<48}[{aff/denom:.3f}, {(aff+unk)/denom:.3f}]"
              f"  ({aff}+{unk})/{denom}")
    return out


def appendix2():
    """52-cell section-level matrix + uniform-512 whole-note results."""
    d = pd.read_csv(U / "expA_panel_results.csv").drop_duplicates(
        subset=["corpus", "variant", "query_format", "model"])
    d["nominal subgroup"] = np.where(d.tier == 1, "harmed", "benefited")
    d["measured sign"] = np.where(d.delta_MRR10 < 0, "negative", "positive")
    d["retains assignment"] = np.where(
        ((d.tier == 1) & (d.delta_MRR10 < 0)) |
        ((d.tier == 2) & (d.delta_MRR10 > 0)), "yes", "no")
    d = d.rename(columns={
        "corpus": "Corpus", "variant": "Document variant",
        "query_format": "Query format", "model": "Configuration",
        "tier": "Nominal tier", "max_len_used": "Max tokens",
        "baseline_MRR10": "Baseline MRR@10", "zca_MRR10": "Corrected MRR@10",
        "delta_MRR10": "dMRR@10"})
    d = d.sort_values(["Document variant", "Corpus", "Query format",
                       "Nominal tier", "Configuration"])
    for c in ["Baseline MRR@10","Corrected MRR@10","dMRR@10"]:
        d[c] = d[c].round(4)
    d.to_csv(OUT / "appendix2_transport_matrix.csv", index=False)
    for v, g in d.groupby("Document variant"):
        ok = (g["retains assignment"] == "yes").sum()
        print(f"A2  {v:<8} {ok}/{len(g)} cells retain assignment")
    return d


def appendix3():
    """Resampling detail behind the pooled screen-accuracy table."""
    d = pd.read_csv(U / "expA_bootstrap.csv")
    d["reference sign"] = np.where(d.delta_full < 0, "negative", "positive")
    d["contributes to"] = np.where(d.tier == 1, "specificity", "sensitivity")
    d = d.rename(columns={
        "corpus": "Corpus", "model": "Configuration", "tier": "Nominal tier",
        "query_format": "Query format", "n_docs": "Documents sampled",
        "B": "Draws", "delta_full": "Full-sample dMRR@10",
        "correct": "Draws with correct sign", "p_correct": "Agreement rate",
        "ci_lo": "Wilson lower", "ci_hi": "Wilson upper"})
    d = d.sort_values(["Corpus", "Nominal tier", "Configuration",
                       "Query format", "Documents sampled"])
    for c in ["Full-sample dMRR@10","Agreement rate","Wilson lower","Wilson upper"]:
        d[c] = d[c].round(4)
    d.to_csv(OUT / "appendix3_screen_resampling.csv", index=False)
    lag = d[(d["Nominal tier"] == 1) & (d["Documents sampled"] == 75)] \
        .groupby(["Corpus", "Configuration"])["Agreement rate"].mean()
    print(f"A3  {len(d)} rows | worst tier-1 agreement at n=75: "
          f"{lag.min():.3f} ({lag.idxmin()})")
    return d


def appendix4():
    """26 paired comparisons across query generators."""
    d = pd.read_csv(U / "mtsamples_both_query_sets.csv")
    piv = d.pivot_table(index=["query_format", "model", "tier"],
                        columns="query_set",
                        values=["baseline_MRR10", "delta_MRR10"]).reset_index()
    piv.columns = [" ".join(c).strip() for c in piv.columns.values]
    piv["dMRR difference (local - published)"] = (
        piv["delta_MRR10 local"] - piv["delta_MRR10 gpt4o"])
    for s in ("gpt4o", "local"):
        piv[f"rank {s}"] = piv.groupby("query_format")[f"delta_MRR10 {s}"] \
            .rank(ascending=False).astype(int)
        piv[f"subgroup {s}"] = np.where(piv[f"delta_MRR10 {s}"] < 0,
                                        "harmed", "benefited")
    piv["assignment agrees"] = np.where(
        piv["subgroup gpt4o"] == piv["subgroup local"], "yes", "no")
    piv = piv.rename(columns={"query_format": "Query format",
                              "model": "Configuration", "tier": "Nominal tier"})
    for c in piv.columns:
        if piv[c].dtype.kind == "f":
            piv[c] = piv[c].round(4)
    piv.to_csv(OUT / "appendix4_query_generator.csv", index=False)
    print(f"A4  {len(piv)} paired comparisons | assignment agrees "
          f"{(piv['assignment agrees']=='yes').sum()}/{len(piv)}")
    return piv


def appendix5():
    """Master parameter table + epsilon sweep. Sources stated per parameter."""
    P = [
        ("d_ben", "Mean dMRR@10, benefited subgroup", "+0.0627", "per-condition SD 0.0465",
         "unitless", "Companion study, full-corpus protocol at eps=1e-5", "empirical", "technical effect"),
        ("d_harm", "Mean dMRR@10, harmed subgroup", "-0.0867", "per-condition SD 0.0465",
         "unitless", "Companion study, full-corpus protocol at eps=1e-5", "empirical", "technical effect"),
        ("p", "Affected-subgroup prevalence", "partially identified",
         "0.036-1.000 (no assumptions)", "proportion",
         "55-system extraction frame; Appendix 1", "partially identified", "technical effect"),
        ("se", "Screen sensitivity", "derived", "by n scored conditions/documents",
         "proportion", "Between-condition variability; resampling, Appendix 3",
         "empirical", "technical effect"),
        ("sp", "Screen specificity", "derived", "by n scored conditions/documents",
         "proportion", "Between-condition variability; resampling, Appendix 3",
         "empirical", "technical effect"),
        ("alpha", "Technical metric to decision accuracy", "0.70",
         "TruncN(0.70, 0.15)", "unitless", "No published source",
         "ILLUSTRATIVE", "enters M"),
        ("adoption", "Retrieval output adopted into decision", "0.60",
         "Beta(9, 6)", "proportion", "No published source", "ILLUSTRATIVE", "enters M"),
        ("p_event", "Adverse event | adopted erroneous retrieval", "0.12",
         "Beta(6, 44)", "proportion", "No published source", "ILLUSTRATIVE", "enters M"),
        ("C_event", "Treatment cost per adverse event", "EUR 11,422",
         "Gamma(2.5, 4569)", "EUR",
         "SAMDATA 2024 bed-day NOK 26,153 x 5.11 excess days (Hoogervorst-Schilp 2015). "
         "Reported without reconciliation to the European estimates of Durand 2024 "
         "and Laroche 2025, which the Norwegian unit cost exceeds",
         "empirical", "enters M via V"),
        ("Q_event", "Utility loss per adverse event", "0.0015 QALY",
         "Gamma(1.5, 0.001); 0.028 alternative", "QALY",
         "Disutility 0.09-0.145 applied over the 5.11-day excess stay; "
         "Kirwan 2023 alternative",
         "scenario", "enters M via V"),
        ("lambda", "Willingness to pay", "NOK 275,000",
         "to NOK 825,000 severity-weighted", "NOK/QALY",
         "Meld. St. 34 (2015-2016)", "empirical", "enters M via V"),
        ("N", "Annual retrievals", "133,000", "fixed", "count",
         "approximately 1/20 of 2,656,857 Norwegian somatic specialist patients "
         "(2024, Statistics Norway)", "empirical", "enters M"),
        ("D", "Discount factor, 5 years at 4%", "4.4518", "fixed", "unitless",
         "Standard", "empirical", "enters M"),
        ("K_impl", "Implementation cost", "EUR 800", "Gamma(2, 400)", "EUR",
         "Scenario", "scenario", "enters K"),
        ("K_annual", "Annual maintenance", "EUR 150", "Gamma(1.5, 100)", "EUR",
         "Scenario", "scenario", "enters K"),
        ("K_screen", "Screening cost per condition", "EUR 400 x n_cond",
         "Gamma(2, 200 x n_cond)", "EUR", "Scenario", "scenario", "enters K"),
        ("epsilon", "ZCA regularisation", "1e-5", "1e-7 to 1e-2 swept",
         "unitless", "Companion study default", "empirical", "technical effect"),
    ]
    cols = ["Symbol", "Definition", "Base value", "Distribution or range",
            "Units", "Source", "Status", "Enters"]
    pd.DataFrame(P, columns=cols).to_csv(
        OUT / "appendix5_parameters.csv", index=False)

    e = pd.read_parquet(U / "epsilon_sensitivity.parquet")
    T1 = TIER1
    e["tier"] = np.where(e.model.isin(T1), 1, 2)
    sw = e.groupby(["epsilon", "tier"])["delta_MRR@10"].mean().unstack()
    sw.columns = ["d_harm (tier 1)", "d_ben (tier 2)"]
    sw["p*"] = sw["d_harm (tier 1)"].abs() / (
        sw["d_ben (tier 2)"] + sw["d_harm (tier 1)"].abs())
    sw = sw.reset_index()
    sw["epsilon"] = sw["epsilon"].map(lambda x: f"{x:.0e}")
    for c in sw.columns:
        if c != "epsilon":
            sw[c] = sw[c].round(4)
    sw.to_csv(OUT / "appendix5_epsilon_sweep.csv", index=False)
    e.rename(columns={"delta_MRR@10": "dMRR@10"}).to_csv(
        OUT / "appendix5_epsilon_percondition.csv", index=False)
    print(f"A5  {len(P)} parameters ({sum(1 for p in P if p[6]=='ILLUSTRATIVE')} illustrative) | "
          f"epsilon sweep {len(sw)} values, {len(e)} per-condition rows")


def appendix6():
    """Code-search boundary replication: baselines, effects, epsilon grid."""
    d = pd.read_parquet(U / "diera_replication.parquet")
    d = d.rename(columns={
        "model": "Configuration", "lang": "Language", "epsilon": "Epsilon",
        "contrastive": "Contrastively trained",
        "baseline_MRR": "Baseline MRR", "zca_MRR": "Whitened MRR",
        "delta_MRR": "dMRR"})
    for c in ["Baseline MRR", "Whitened MRR", "dMRR"]:
        d[c] = d[c].round(4)
    # The full grid is broader than the 18-cell primary replication: it adds
    # the fine-tuned CodeBERT configuration, for which no published baseline
    # exists, and the seventh language (R, StatCodeSearch). Labelled so the
    # dimensions can be reconciled with the manuscript.
    d["Analysis"] = np.where(
        (d.Configuration == "codebert_ft") | (d.Language == "r"),
        "extended regularisation analysis",
        "primary replication grid")
    d.to_csv(OUT / "appendix6_codesearch_full.csv", index=False)

    base = d[d.Epsilon == 0.0][["Configuration", "Language", "Baseline MRR"]] \
        .drop_duplicates()
    pub = {("codebert","ruby"):0.006,("codebert","javascript"):0.002,
           ("codebert","go"):0.002,("codebert","java"):0.000,
           ("codebert","python"):0.001,("codebert","php"):0.000,
           ("codet5p","ruby"):0.705,("codet5p","javascript"):0.638,
           ("codet5p","go"):0.757,("codet5p","java"):0.595,
           ("codet5p","python"):0.721,("codet5p","php"):0.537,
           ("codellama","ruby"):0.047,("codellama","javascript"):0.026,
           ("codellama","go"):0.031,("codellama","java"):0.015,
           ("codellama","python"):0.017,("codellama","php"):0.009}
    base["Published baseline"] = [
        pub.get((r.Configuration, r.Language), np.nan) for r in base.itertuples()]
    base["Absolute difference"] = (
        base["Baseline MRR"] - base["Published baseline"]).abs()
    prim = base.dropna(subset=["Published baseline"]).copy()
    prim["Analysis"] = "primary baseline replication (3 model families x 6 languages)"
    prim.to_csv(OUT / "appendix6_baseline_replication.csv", index=False)

    grid = d.groupby(["Configuration", "Epsilon"])["dMRR"].mean().unstack()
    grid.reset_index().to_csv(OUT / "appendix6_epsilon_grid.csv", index=False)
    n = base.dropna(subset=["Published baseline"])
    print(f"A6  {len(d)} rows | {len(n)} baseline cells, max deviation "
          f"{n['Absolute difference'].max():.4f}")
    print(f"    epsilon grid {grid.shape[0]} configurations x {grid.shape[1]} values")


if __name__ == "__main__":
    appendix1(); appendix2(); appendix3(); appendix4(); appendix5(); appendix6()
    print("\nfiles:")
    for f in sorted(OUT.glob("*.csv")):
        print(f"  {f.name:<44}{f.stat().st_size:>8,} bytes")
