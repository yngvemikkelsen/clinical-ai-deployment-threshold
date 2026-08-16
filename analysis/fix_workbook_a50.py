#!/usr/bin/env python3
"""
Correct the A50 audit note in the extraction workbook.

WHY
---
The workbook records the ORIGINAL coding rationale for A50 (Shi /
Chat-Orthopedist):

    "No retrieval/similarity-tuned checkpoint is identified. Recoded as
     ambiguous in sensitivity notes."

The manuscript establishes that this rationale is invalid. Citing the
base-model paper is the standard citation practice for the contrastively
fine-tuned sentence-transformer checkpoint, so "no similarity-tuned checkpoint
is identified" does not follow from the citation pattern.

The valid basis is a pooling mismatch, read from the primary article's methods:
Section 3.2 of doi:10.1145/3584371.3612956 states that the representation is
taken at the [CLS] token, while all-mpnet-base-v2 applies its contrastive
objective to MEAN-POOLED output. The representation actually used for retrieval
is therefore not the one the objective calibrated, and membership follows from
the extracted representation rather than from the training label.

The Tier 2 coding is unchanged. Only the stated reason changes — but the reason
is what the manuscript's pooling argument rests on, and A50 is one of only two
confirmed affected systems and the only inferable extraction step in the frame.
Depositing the workbook with the superseded rationale would put the supporting
data in contradiction with the paper on its load-bearing case.

WHAT THIS DOES
--------------
Rewrites the A50 audit note, adds a provenance column recording the superseded
rationale and the correction, and leaves every other cell untouched. The
original file is not modified; a corrected copy is written alongside it.

    python fix_workbook_a50.py --in  p_tier2_wide_frame_extraction_audit.xlsx \\
                               --out p_tier2_wide_frame_extraction_audit_v2.xlsx
"""
from __future__ import annotations

import argparse
import shutil
from pathlib import Path

from openpyxl import load_workbook

TARGET_ID = "A50"

NEW_NOTE = (
    "Tier 2 on the basis of the representation actually extracted. The primary "
    "article reports extraction of the [CLS] representation (Section 3.2, "
    "doi:10.1145/3584371.3612956) and identifies MPNet by reference to the "
    "original MPNet model (Song et al.) rather than to a retrieval-trained "
    "sentence-embedding checkpoint. On the reported implementation the "
    "retrieval representation was therefore not shown to have been calibrated "
    "by a retrieval-specific objective. Membership follows from the extracted "
    "representation, not from the training label."
)

SUPERSEDED = (
    "SUPERSEDED RATIONALE: 'No retrieval/similarity-tuned checkpoint is "
    "identified. Recoded as ambiguous in sensitivity notes.' This basis is "
    "invalid: citing the base-model paper is the standard citation for the "
    "contrastively fine-tuned sentence-transformer checkpoint, so the absence "
    "of a named similarity-tuned checkpoint does not follow. Tier 2 coding is "
    "unchanged; only the basis is corrected."
)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="src", type=Path,
                    default=Path("p_tier2_wide_frame_extraction_audit.xlsx"))
    ap.add_argument("--out", dest="dst", type=Path,
                    default=Path("p_tier2_wide_frame_extraction_audit_v2.xlsx"))
    a = ap.parse_args()
    if not a.src.exists():
        raise SystemExit(f"not found: {a.src}")

    shutil.copy(a.src, a.dst)
    wb = load_workbook(a.dst)
    ws = wb["Analysis_Set"]

    hdr = [c.value for c in ws[1]]
    try:
        c_id = hdr.index("Canonical ID") + 1
        c_note = hdr.index("Audit note") + 1
    except ValueError as e:
        raise SystemExit(f"expected column missing: {e}")

    # add the provenance column rather than overwriting history
    c_prov = len(hdr) + 1
    ws.cell(row=1, column=c_prov, value="Coding provenance")

    hit = 0
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=c_id).value).strip() == TARGET_ID:
            old = ws.cell(row=r, column=c_note).value
            ws.cell(row=r, column=c_note, value=NEW_NOTE)
            ws.cell(row=r, column=c_prov, value=SUPERSEDED)
            hit += 1
            print(f"row {r}  {TARGET_ID}")
            print(f"  was: {old}")
            print(f"  now: {NEW_NOTE[:78]}...")

    if hit != 1:
        raise SystemExit(f"expected exactly one {TARGET_ID} row, found {hit} "
                         "- nothing written")

    wb.save(a.dst)
    print(f"\n-> {a.dst}")
    print("Tier codings unchanged in all frames; only the stated basis and a "
          "new provenance column differ.")

    # verify the tier columns really are untouched
    wb0 = load_workbook(a.src, read_only=True, data_only=True)
    w0, w1 = wb0["Analysis_Set"], load_workbook(a.dst, read_only=True,
                                                data_only=True)["Analysis_Set"]
    r0 = [list(r) for r in w0.iter_rows(values_only=True)]
    r1 = [list(r) for r in w1.iter_rows(values_only=True)]
    tier_cols = [hdr.index(c) for c in
                 ("Core tier", "Expanded tier", "Adversarial tier",
                  "Core include", "Expanded include", "Adversarial include")]
    diffs = sum(1 for a_, b_ in zip(r0[1:], r1[1:])
                for i in tier_cols if a_[i] != b_[i])
    print(f"verification: {diffs} differences across tier and include columns "
          f"(expected 0)")


if __name__ == "__main__":
    main()
